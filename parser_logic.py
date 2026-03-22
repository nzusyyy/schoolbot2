import openpyxl
import os
import logging
import shutil
from zipfile import ZipFile
import re
import time
import glob
import uuid
from datetime import datetime

DATA_DIR = "data"

def get_day_name(date_str: str):
    """Превращает дату в '12.03 (Четверг)'"""
    try:
        # Пытаемся распарсить дату (ДД.ММ.ГГГГ)
        date_obj = datetime.strptime(date_str, "%d.%m.%p") if len(date_str) < 6 else datetime.strptime(date_str, "%d.%m.%Y")
    except:
        try:
            date_obj = datetime.strptime(date_str, "%d.%m")
            date_obj = date_obj.replace(year=datetime.now().year)
        except:
            return date_str

    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    return f"{date_str} ({days[date_obj.weekday()]})"

def fix_excel_format(file_path: str):
    unique_id = str(uuid.uuid4())[:8]
    tmp_folder = f"temp_fix_{unique_id}"
    temp_zip_name = f"fixed_temp_{unique_id}"
    try:
        os.makedirs(tmp_folder, exist_ok=True)
        with ZipFile(file_path, 'r') as excel_container:
            excel_container.extractall(tmp_folder)
        xl_dir = os.path.join(tmp_folder, "xl")
        if os.path.exists(xl_dir):
            actual_files = os.listdir(xl_dir)
            if "SharedStrings.xml" in actual_files:
                old_path = os.path.join(xl_dir, "SharedStrings.xml")
                temp_path = os.path.join(xl_dir, "SharedStrings.xml.tmp")
                new_path = os.path.join(xl_dir, "sharedStrings.xml")
                os.rename(old_path, temp_path)
                os.rename(temp_path, new_path)
                shutil.make_archive(temp_zip_name, "zip", tmp_folder)
                time.sleep(0.5)
                if os.path.exists(f"{temp_zip_name}.zip"):
                    if os.path.exists(file_path): os.remove(file_path)
                    shutil.move(f"{temp_zip_name}.zip", file_path)
        return True
    except Exception as e:
        logging.error(f"Ошибка фикса: {e}")
        return False
    finally:
        shutil.rmtree(tmp_folder, ignore_errors=True)
        if os.path.exists(f"{temp_zip_name}.zip"): os.remove(f"{temp_zip_name}.zip")

def get_menu_from_file(file_path: str, target_date: str):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 1. Ищем колонку с датой (проверяем первые 5 строк на всякий случай)
        target_col_idx = None
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(row=r, column=c).value or "")
                if target_date in cell_val:
                    target_col_idx = c
                    break
            if target_col_idx: break
        
        if not target_col_idx:
            wb.close()
            return None

        # 2. Ищем колонку с блюдами и стартовую строку данных
        dish_col_idx = 4 # По умолчанию 4-я
        data_start_row = 3
        found_header = False
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                if cell_val == "блюдо":
                    dish_col_idx = c
                    data_start_row = r + 1
                    found_header = True
                    break
            if found_header: break

        menu_data = {}
        current_type = None
        
        # 3. Читаем данные
        for row_idx in range(data_start_row, ws.max_row + 1):
            meal_type = str(ws.cell(row=row_idx, column=1).value or "").strip()
            dish_name = str(ws.cell(row=row_idx, column=dish_col_idx).value or "").strip()
            
            # Если в первой колонке что-то есть, это новый прием пищи (Завтрак, Обед...)
            if meal_type and meal_type.lower() != 'none':
                current_type = meal_type
                if current_type not in menu_data:
                    menu_data[current_type] = []
            
            # Если есть название блюда, добавляем его в текущий тип приема пищи
            if current_type and dish_name and dish_name.lower() != 'none':
                menu_data[current_type].append(f"  ▫️ {dish_name}")

        wb.close()
        if not menu_data: return None

        full_date_name = get_day_name(target_date)
        msg = f"✨ *ШКОЛЬНОЕ МЕНЮ* ✨\n"
        msg += f"📅 `{full_date_name}`\n"
        msg += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n\n"
        
        icons = {"завтрак": "🥣", "обед": "🍲", "полдник": "🍪", "ужин": "🌙"}
        for m_type, dishes in menu_data.items():
            if not dishes: continue
            icon = icons.get(m_type.lower(), "🍴")
            msg += f"{icon} *{m_type.upper()}*\n"
            msg += "\n".join(dishes) + "\n\n"
        
        msg += "⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯⎯\n"
        msg += "💌 _Приятного аппетита!_"
        return msg
    except Exception as e:
        logging.error(f"Error parsing {file_path}: {e}")
        return None

def get_menu_by_date(target_date: str):
    if not os.path.exists(DATA_DIR): return "🤷 Файлы меню не загружены."
    files = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    for file in files:
        result = get_menu_from_file(file, target_date)
        if result: return result
    return f"🤷 На дату {target_date} меню не найдено."

def get_available_dates():
    if not os.path.exists(DATA_DIR): return []
    all_dates = set()
    files = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    date_pattern = re.compile(r'\d{2}\.\d{2}\.\d{4}')
    for file in files:
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            # Проверяем только первую строку (заголовки)
            for col_idx in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=1, column=col_idx).value or "")
                match = date_pattern.search(cell_value)
                if match:
                    all_dates.add(match.group())
            wb.close()
        except: continue
    
    # Сортировка дат
    def sort_key(d):
        try: return datetime.strptime(d, "%d.%m.%Y")
        except: return datetime.now()
        
    return sorted(list(all_dates), key=sort_key)

def delete_menu_by_date(target_date: str):
    """Находит и удаляет файлы, содержащие указанную дату"""
    if not os.path.exists(DATA_DIR): return False
    
    files = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    deleted_any = False
    date_pattern = re.compile(re.escape(target_date))
    
    for file_path in files:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            found = False
            for col_idx in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=1, column=col_idx).value or "")
                if date_pattern.search(cell_value):
                    found = True
                    break
            wb.close()
            
            if found:
                os.remove(file_path)
                deleted_any = True
        except Exception as e:
            logging.error(f"Error deleting date {target_date} from {file_path}: {e}")
            continue
            
    return deleted_any
